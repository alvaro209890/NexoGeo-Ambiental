# -*- coding: utf-8 -*-
"""core.xlsx_builder — estilo padrão das planilhas (.xlsx).

Genérico: cores institucionais, bordas, cabeçalho azul, linha de total verde e
larguras de coluna — extraído dos scripts de quantitativos (onde estava copiado
em 3–4 arquivos). Nenhuma planilha específica de imóvel aqui.
"""
from __future__ import annotations

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Cores institucionais (ARGB)
AZUL = "FF1F4E79"
VERDE = "FF548235"
CINZA = "FFF2F2F2"


def fill(cor_argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=cor_argb)


# Fontes / alinhamentos reutilizáveis
BRANCO_BOLD = Font(color="FFFFFFFF", bold=True, size=11)
TITULO = Font(color="FFFFFFFF", bold=True, size=14)
BOLD = Font(bold=True)
NOTA = Font(italic=True, size=9, color="FF595959")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

_thin = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FMT_HA = "#,##0.0000"
FMT_PCT = '0.00"%"'


def titulo_merge(ws, intervalo: str, texto: str, fundo: str = AZUL, altura: int = 30):
    """Faz merge e escreve um título estilizado (ex.: 'A1:F1')."""
    ws.merge_cells(intervalo)
    cel = ws[intervalo.split(":")[0]]
    cel.value = texto
    cel.font = TITULO
    cel.fill = fill(fundo)
    cel.alignment = CENTER
    ws.row_dimensions[cel.row].height = altura
    return cel


def nota_merge(ws, intervalo: str, texto: str, altura: int = 26):
    """Linha de fonte/observação (itálico cinza)."""
    ws.merge_cells(intervalo)
    cel = ws[intervalo.split(":")[0]]
    cel.value = texto
    cel.font = NOTA
    cel.alignment = LEFT
    ws.row_dimensions[cel.row].height = altura
    return cel


def cabecalho(ws, linha: int, headers: list[str], fundo: str = AZUL, altura: int = 30):
    """Escreve a linha de cabeçalho (fundo azul, texto branco, bordas)."""
    for c, h in enumerate(headers, start=1):
        cel = ws.cell(row=linha, column=c, value=h)
        cel.font = BRANCO_BOLD
        cel.fill = fill(fundo)
        cel.alignment = CENTER
        cel.border = BORDER
    ws.row_dimensions[linha].height = altura


def larguras(ws, widths: list[int]):
    """Define larguras das colunas, na ordem."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def celula(ws, linha: int, col: int, valor, *, align=None, fmt=None, zebra=False,
           fonte=None, fundo=None):
    """Escreve uma célula de corpo com borda e formatação opcional."""
    cel = ws.cell(row=linha, column=col, value=valor)
    cel.border = BORDER
    if align is not None:
        cel.alignment = align
    if fmt:
        cel.number_format = fmt
    if fonte is not None:
        cel.font = fonte
    if fundo:
        cel.fill = fill(fundo)
    elif zebra and linha % 2 == 0:
        cel.fill = fill(CINZA)
    return cel
