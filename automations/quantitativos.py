# -*- coding: utf-8 -*-
"""Automação: quantitativo de área cultivável por CAR -> .xlsx.

Reproduz o antigo ``gerar_quantitativos.py`` (3 abas: Quantitativos, Detalhe da
Área Cultivável, Detalhe CAR), agora **genérico** — shapes, fazendas, CRS e campos
do .dbf vêm do ``Projeto``. Nenhum nome de fazenda escrito no código.

Uso (a partir de software/):
    python -m automations.quantitativos <projeto.json> [saida.xlsx]
"""
from __future__ import annotations

import os
import openpyxl

from core import geo, io
from core import xlsx_builder as xb

NOME_PADRAO = "Quantitativos_CAR_Area_Cultivavel.xlsx"


def _area_car(projeto, car) -> float:
    try:
        return float(car.attr(projeto, "area_ha"))
    except (TypeError, ValueError):
        return 0.0


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def _aba_quant(ws, projeto, cars, assoc):
    ws.title = "Quantitativos"
    m = projeto.municipio
    xb.titulo_merge(ws, "A1:F1",
                    f"QUANTITATIVO DE ÁREAS POR CAR — {projeto.imovel} "
                    f"(Município {m.nome}/{m.uf} - {m.ibge})")
    xb.nota_merge(ws, "A2:F2",
                  "Fonte: shapes CAR_ATP (SEMA-MT). Área cultivável: tabela de atributos "
                  "do shape de área cultivável; atribuição por sobreposição > 50%.")
    hr = 4
    xb.cabecalho(ws, hr, ["#", "CAR / Fazenda", "Área Total (ha)",
                          "Área Cultivável (ha)", "% Cultivável", "Nº polígonos cult."])
    r = hr + 1
    tot_area = tot_cult = 0.0
    tot_n = 0
    for idx, c in enumerate(cars, start=1):
        a_car = _area_car(projeto, c)
        a_cult = assoc[c.id]["area"]
        n = assoc[c.id]["n"]
        pct = (a_cult / a_car * 100) if a_car else 0
        tot_area += a_car
        tot_cult += a_cult
        tot_n += n
        xb.celula(ws, r, 1, idx, align=xb.CENTER, zebra=True)
        xb.celula(ws, r, 2, c.nome, align=xb.LEFT, zebra=True)
        xb.celula(ws, r, 3, round(a_car, 4), align=xb.RIGHT, fmt=xb.FMT_HA, zebra=True)
        xb.celula(ws, r, 4, round(a_cult, 4), align=xb.RIGHT, fmt=xb.FMT_HA, zebra=True)
        xb.celula(ws, r, 5, round(pct, 2), align=xb.CENTER, fmt=xb.FMT_PCT, zebra=True)
        xb.celula(ws, r, 6, n, align=xb.CENTER, zebra=True)
        r += 1
    pct_tot = (tot_cult / tot_area * 100) if tot_area else 0
    totais = [("", xb.CENTER, None), ("TOTAL", xb.LEFT, None),
              (round(tot_area, 4), xb.RIGHT, xb.FMT_HA), (round(tot_cult, 4), xb.RIGHT, xb.FMT_HA),
              (round(pct_tot, 2), xb.CENTER, xb.FMT_PCT), (int(tot_n), xb.CENTER, None)]
    for col, (v, al, fmt) in enumerate(totais, start=1):
        xb.celula(ws, r, col, v, align=al, fmt=fmt, fonte=xb.BRANCO_BOLD, fundo=xb.VERDE)
    xb.larguras(ws, [5, 30, 16, 18, 12, 14])
    ws.freeze_panes = "A5"


def _aba_detalhe(ws, cars, detalhe):
    nome_by_id = {c.id: c.nome for c in cars}
    xb.titulo_merge(ws, "A1:C1", "DETALHE — Polígonos de Área Cultivável",
                    fundo=xb.VERDE, altura=28)
    xb.cabecalho(ws, 3, ["Polígono #", "Área Cultivável (ha)", "CAR / Fazenda"], fundo=xb.VERDE)
    rr = 4
    for i, area, cid in detalhe:
        xb.celula(ws, rr, 1, i, align=xb.CENTER)
        xb.celula(ws, rr, 2, round(area, 4), align=xb.RIGHT, fmt=xb.FMT_HA)
        xb.celula(ws, rr, 3, nome_by_id.get(cid, "NÃO ATRIBUÍDO"), align=xb.LEFT)
        rr += 1
    cel = ws.cell(row=rr, column=1, value="TOTAL")
    cel.font = xb.BOLD
    tc = ws.cell(row=rr, column=2, value=round(sum(d[1] for d in detalhe), 4))
    tc.font = xb.BOLD
    tc.number_format = xb.FMT_HA
    tc.alignment = xb.RIGHT
    xb.larguras(ws, [12, 22, 30])
    ws.freeze_panes = "A4"


def _aba_car(ws, projeto, cars, assoc):
    xb.cabecalho(ws, 1, ["CAR / Fazenda", "Código CAR Federal", "Módulos Fiscais",
                         "Área Total (ha)", "Área Cultivável (ha)", "% Cultivável", "Proprietários"])
    rr = 2
    for c in cars:
        a_car = _area_car(projeto, c)
        a_cult = assoc[c.id]["area"]
        pct = round(a_cult / a_car * 100, 2) if a_car else 0
        xb.celula(ws, rr, 1, c.nome, align=xb.LEFT)
        xb.celula(ws, rr, 2, c.attr(projeto, "car_federal", ""), align=xb.LEFT)
        xb.celula(ws, rr, 3, _num(c.attr(projeto, "modulos")), align=xb.CENTER)
        xb.celula(ws, rr, 4, round(a_car, 4), align=xb.RIGHT, fmt=xb.FMT_HA)
        xb.celula(ws, rr, 5, round(a_cult, 4), align=xb.RIGHT, fmt=xb.FMT_HA)
        xb.celula(ws, rr, 6, pct, align=xb.CENTER, fmt=xb.FMT_PCT)
        xb.celula(ws, rr, 7, c.attr(projeto, "proprietarios", ""), align=xb.LEFT)
        rr += 1
    xb.larguras(ws, [28, 42, 14, 16, 18, 12, 50])
    ws.freeze_panes = "A2"


def gerar(projeto, destino: str | None = None) -> str:
    """Gera a planilha de quantitativos de área cultivável e devolve o caminho salvo."""
    cars = geo.carregar_cars(projeto)
    nome_cult = projeto.quantitativos.get("shapes", {}).get("cultivavel", "Area_Cultivavel")
    shp_cult = os.path.join(projeto.caminho("shapes"), nome_cult + ".shp")
    if not os.path.exists(shp_cult):
        raise FileNotFoundError(f"shape de área cultivável não encontrado: {shp_cult}")
    assoc, detalhe = geo.atribuir(cars, shp_cult, projeto.crs_utm)

    wb = openpyxl.Workbook()
    _aba_quant(wb.active, projeto, cars, assoc)
    _aba_detalhe(wb.create_sheet("Detalhe Area Cultivavel"), cars, detalhe)
    _aba_car(wb.create_sheet("Detalhe CAR"), projeto, cars, assoc)

    destino = destino or io.caminho_resultado(projeto, NOME_PADRAO)
    wb.save(destino)
    return destino


def _main(argv: list[str]) -> int:
    from core.config import load_projeto
    if len(argv) < 2:
        print("uso: python -m automations.quantitativos <projeto.json> [saida.xlsx]")
        return 2
    proj = load_projeto(argv[1])
    out = gerar(proj, argv[2] if len(argv) > 2 else None)
    print("Gerado:", out)
    return 0


if __name__ == "__main__":
    import sys
    raise SystemExit(_main(sys.argv))
